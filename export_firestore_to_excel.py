import re
import json
from datetime import datetime

import pandas as pd
from firebase_admin import credentials, firestore, initialize_app
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

SERVICE_ACCOUNT_JSON = "serviceAccountKey.json"
OUTPUT_XLSX = "sessions_one_document_per_sheet_polished.xlsx"


def safe_cell_value(value):
    """Convert Firestore values into Excel-friendly values."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)

    return value


def question_sort_key(doc):
    """Sort question docs numerically if ids are like 1, 2, 3..."""
    try:
        return int(doc.id)
    except ValueError:
        return doc.id


def make_valid_sheet_name(name, used_names):
    """
    Make the sheet name valid for Excel.
    Rules:
    - max length 31
    - cannot contain: : \\ / ? * [ ]
    - must be unique
    """
    cleaned = re.sub(r'[:\\/*?\[\]]', "_", name).strip()
    if not cleaned:
        cleaned = "Sheet"

    cleaned = cleaned[:31]
    base_name = cleaned
    counter = 1

    while cleaned in used_names:
        suffix = f"_{counter}"
        cleaned = base_name[:31 - len(suffix)] + suffix
        counter += 1

    used_names.add(cleaned)
    return cleaned


def apply_table_header_style(ws, row_number):
    """Apply styling to a header row."""
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")  # white text
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_data_borders(ws, start_row, end_row, start_col, end_col):
    """Apply light borders to a rectangular range."""
    thin_border = Border(
        left=Side(style="thin", color="E6E6E6"),
        right=Side(style="thin", color="E6E6E6"),
        top=Side(style="thin", color="E6E6E6"),
        bottom=Side(style="thin", color="E6E6E6"),
    )

    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def auto_fit_columns(ws):
    """Set column widths based on the longest cell value."""
    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        column_letter = get_column_letter(column_index)

        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        adjusted_width = min(max_length + 2, 50)  # cap width so it doesn't get too huge
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)


def style_session_key_value_section(ws, start_row, end_row):
    """Style the top key-value session details section."""
    label_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    label_font = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for row in range(start_row, end_row + 1):
        # Column A = Field
        field_cell = ws.cell(row=row, column=1)
        value_cell = ws.cell(row=row, column=2)

        field_cell.fill = label_fill
        field_cell.font = label_font
        field_cell.border = thin_border
        field_cell.alignment = Alignment(vertical="top", wrap_text=True)

        value_cell.border = thin_border
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    initialize_app(credentials.Certificate(SERVICE_ACCOUNT_JSON))
    db = firestore.client()

    session_docs = list(db.collection("sessions").stream())
    used_sheet_names = set()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for session_doc in session_docs:
            session_id = session_doc.id
            session_data = session_doc.to_dict() or {}

            sheet_name = make_valid_sheet_name(session_id, used_sheet_names)

            # -----------------------------
            # Build session details section
            # -----------------------------
            session_rows = [{"Field": "sessionId", "Value": session_id}]
            for key, value in session_data.items():
                session_rows.append({
                    "Field": key,
                    "Value": safe_cell_value(value),
                })

            session_df = pd.DataFrame(session_rows)

            # -----------------------------
            # Build questions section
            # -----------------------------
            question_docs = list(session_doc.reference.collection("questions").stream())
            question_docs.sort(key=question_sort_key)

            question_rows = []
            for question_doc in question_docs:
                question_data = question_doc.to_dict() or {}
                row = {"questionId": question_doc.id}

                for key, value in question_data.items():
                    row[key] = safe_cell_value(value)

                question_rows.append(row)

            questions_df = pd.DataFrame(question_rows)

            # -----------------------------
            # Write session details
            # -----------------------------
            session_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=0,
            )

            # Row positions in Excel are 1-based in openpyxl
            session_header_row = 1
            session_data_start_row = 2
            session_data_end_row = len(session_df) + 1

            # -----------------------------
            # Add questions title
            # -----------------------------
            questions_title_row = len(session_df) + 4

            # -----------------------------
            # Write questions table
            # -----------------------------
            questions_start_row_zero_based = questions_title_row  # pandas uses 0-based
            questions_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=questions_start_row_zero_based,
            )

            # Get worksheet for styling
            ws = writer.book[sheet_name]

            # -----------------------------
            # Style session section
            # -----------------------------
            apply_table_header_style(ws, session_header_row)
            style_session_key_value_section(ws, session_data_start_row, session_data_end_row)

            # -----------------------------
            # Questions section title
            # -----------------------------
            title_cell = ws.cell(row=questions_title_row, column=1)
            title_cell.value = "Questions"
            title_cell.font = Font(bold=True, size=13, color="000000")
            title_cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

            # -----------------------------
            # Style questions header/data
            # -----------------------------
            questions_header_row = questions_title_row + 1
            apply_table_header_style(ws, questions_header_row)

            if not questions_df.empty:
                questions_data_start_row = questions_header_row + 1
                questions_data_end_row = questions_header_row + len(questions_df)
                questions_end_col = len(questions_df.columns)

                apply_data_borders(
                    ws,
                    start_row=questions_data_start_row,
                    end_row=questions_data_end_row,
                    start_col=1,
                    end_col=questions_end_col,
                )

                # Freeze pane below the questions header row
                ws.freeze_panes = f"A{questions_data_start_row}"
            else:
                ws.freeze_panes = "A2"

            # -----------------------------
            # Auto-fit columns
            # -----------------------------
            auto_fit_columns(ws)

    print(f"Export completed successfully: {OUTPUT_XLSX}")
    print(f"Total session sheets created: {len(session_docs)}")


if __name__ == "__main__":
    main()